"""
Create a sample pega_accounts.xlsx with dummy Partner and Customer records.
Run: python data/create_sample_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PARTNERS = [
    "Accenture", "Infosys BPM", "Wipro", "Cognizant", "TCS", "HCL Technologies",
    "Capgemini", "Deloitte", "EY", "PwC", "IBM", "DXC Technology",
]

CUSTOMERS = [
    "Tokio Marine HCC", "Bank of America", "JPMorgan Chase", "Wells Fargo",
    "Lloyds Banking Group", "HSBC", "Deutsche Bank", "Citibank",
    "Autodesk", "Adobe Systems", "Salesforce", "SAP",
    "United Health Group", "Aetna", "MetLife", "Prudential Financial",
    "Delta Air Lines", "American Airlines", "FedEx", "UPS",
    "Walmart", "Target Corporation", "Costco Wholesale",
    "General Electric", "Siemens", "Honeywell", "3M",
    "Anthem", "Humana", "CVS Health",
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pega Accounts"

# Header
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")

ws["A1"] = "Company Name"
ws["B1"] = "Type"
ws["C1"] = "Notes"

for col_letter in ["A", "B", "C"]:
    cell = ws[f"{col_letter}1"]
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 40

partner_fill = PatternFill(fill_type="solid", fgColor="FFF3CD")
customer_fill = PatternFill(fill_type="solid", fgColor="D1FAE5")

row = 2
for name in PARTNERS:
    ws.cell(row=row, column=1, value=name).fill = partner_fill
    ws.cell(row=row, column=2, value="Partner").fill = partner_fill
    ws.cell(row=row, column=3, value="Pega Partner — do not research").fill = partner_fill
    row += 1

for name in CUSTOMERS:
    ws.cell(row=row, column=1, value=name).fill = customer_fill
    ws.cell(row=row, column=2, value="Customer").fill = customer_fill
    ws.cell(row=row, column=3, value="Known Pega customer").fill = customer_fill
    row += 1

ws.freeze_panes = "A2"
wb.save("data/pega_accounts.xlsx")
print(f"✅ Created data/pega_accounts.xlsx with {len(PARTNERS)} partners and {len(CUSTOMERS)} customers.")

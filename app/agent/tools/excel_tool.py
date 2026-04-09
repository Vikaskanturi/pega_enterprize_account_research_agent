"""
Excel tool — reads the Pega reference file and writes the 33-column output.
"""
import os
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── Reference Excel (Partner/Customer list) ───────────────────────────────────

def load_pega_accounts(file_path: Optional[str] = None) -> dict[str, str]:
    """
    Load the Pega partner/customer reference Excel file.
    Returns a dict: {normalized_company_name: 'Customer' | 'Partner'}
    """
    path = file_path or os.getenv("PEGA_ACCOUNTS_FILE", "data/pega_accounts.xlsx")
    if not os.path.exists(path):
        return {}

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    accounts = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        company = str(row[0]).strip()
        category = str(row[1]).strip() if len(row) > 1 and row[1] else "Customer"
        accounts[_normalize(company)] = category

    wb.close()
    return accounts


def classify_company(company_name: str, accounts: dict[str, str]) -> Optional[str]:
    """
    Look up a company in the reference dict.
    Returns 'Customer', 'Partner', or None if not found.
    Handles name variations (abbreviations, Inc., Ltd., etc.)
    """
    normalized = _normalize(company_name)

    # Exact match
    if normalized in accounts:
        return accounts[normalized]

    # Fuzzy match — strip common suffixes and try again
    stripped = _strip_suffixes(normalized)
    for key, val in accounts.items():
        if _strip_suffixes(key) == stripped:
            return val

    # Partial match — company name is a substring
    for key, val in accounts.items():
        if stripped in key or key in stripped:
            return val

    return None


def _normalize(name: str) -> str:
    return name.lower().strip()


def _strip_suffixes(name: str) -> str:
    suffixes = ["inc", "inc.", "ltd", "ltd.", "llc", "llc.", "corp", "corp.",
                "limited", "group", "holdings", "international", "technologies",
                "technology", "solutions", "services", "systems"]
    parts = name.split()
    while parts and parts[-1].lower().rstrip(".") in suffixes:
        parts.pop()
    return " ".join(parts)


# ── Output Excel (33 columns) ─────────────────────────────────────────────────

ENTERPRISE_COLORS = {
    "E1": "FF4444",
    "E1.1": "FF8C00",
    "E2": "4A90D9",
    "E3": "22C55E",
}

HEADER_BG = "1E293B"
HEADER_FG = "F8FAFC"
ROW_ALT = "F1F5F9"
ROW_NORMAL = "FFFFFF"


def write_output_excel(rows: list[dict], output_path: str = "output/research_results.xlsx"):
    """
    Write research results to a formatted 33-column Excel file.
    Each row is a dict matching ResearchState.to_excel_row().
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Research Results"

    if not rows:
        wb.save(output_path)
        return output_path

    headers = list(rows[0].keys())

    # ── Header row ────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=10)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 40

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, 2):
        is_alt = row_idx % 2 == 0
        bg = ROW_ALT if is_alt else ROW_NORMAL
        enterprise_type = row_data.get("Enterprise Type", "")
        etype_color = ENTERPRISE_COLORS.get(enterprise_type)

        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

            # Highlight Enterprise Type column
            if header == "Enterprise Type" and etype_color:
                cell.fill = PatternFill(fill_type="solid", fgColor=etype_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            # Highlight Pega Usage column
            elif header == "Pega Usage Confirmed":
                if value == "Yes":
                    cell.fill = PatternFill(fill_type="solid", fgColor="D1FAE5")
                elif value == "No":
                    cell.fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
                elif value == "Unconfirmed":
                    cell.fill = PatternFill(fill_type="solid", fgColor="FEF9C3")
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)

        ws.row_dimensions[row_idx].height = 30

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "Company Name": 25,
        "Parent Company": 25,
        "India Subsidiary": 25,
        "Industry": 20,
        "Headquarters Location": 22,
        "Primary Revenue Source": 25,
        "Software or Non-Software": 18,
        "Annual Revenue (USD)": 18,
        "Pega Customer / Partner": 18,
        "Pega Usage Confirmed": 18,
        "Pega Evidence": 40,
        "Subsidiaries & Associated Companies": 40,
        "GCCs in India": 12,
        "Number of GCCs": 14,
        "GCC Locations": 30,
        "Main GCC in India": 25,
        "Software Development Model": 20,
        "Signals of Service Companies": 20,
        "Service Companies Identified": 30,
        "Hiring for Tech Roles": 16,
        "Tech Roles Being Hired For": 35,
        "Total Employee Count (Org-wide)": 22,
        "Employee Count (India)": 20,
        "Engineering Count (Org-wide)": 22,
        "Engineering Count (India)": 22,
        "IT Count (Org-wide)": 18,
        "IT Count (India)": 18,
        "SDET & QA Count (Org-wide)": 22,
        "SDET & QA Count (India)": 22,
        "Engineering % of Total Headcount": 24,
        "Other Enterprise Platforms": 35,
        "Enterprise Type": 14,
        "Research Notes / Comments": 50,
    }

    for col_idx, header in enumerate(headers, 1):
        width = col_widths.get(header, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def append_to_master_excel(row_data: dict, output_path: str = "pega_master_data.xlsx"):
    """
    Append a research result row to a master Excel file.
    If the file doesn't exist, it creates a new one.
    Deprecated: prefer upsert_to_master_excel for the shared output file.
    """
    return upsert_to_master_excel(row_data, output_path=output_path)


MASTER_OUTPUT_PATH = "output/research_results.xlsx"


def upsert_to_master_excel(
    row_data: dict,
    output_path: str = MASTER_OUTPUT_PATH,
) -> str:
    """
    Save a research result into a single shared Excel file.

    Behaviour:
    - If the file does not exist → create it fresh (with header + this row).
    - If the file exists and a row already has the same Company Name
      (case-insensitive) → update every column in that row in-place.
    - If the file exists but no matching company row found → append a new row.

    Args:
        row_data: Dict matching ResearchState.to_excel_row().
        output_path: Path to the shared master Excel file.

    Returns:
        The output_path that was written.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # ── File does not exist yet → create brand-new ────────────────────────────
    if not os.path.exists(output_path):
        return write_output_excel([row_data], output_path=output_path)

    # ── File exists → open and upsert ─────────────────────────────────────────
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Read existing headers from row 1
    headers = [cell.value for cell in ws[1]]

    # Determine the column index for "Company Name"
    try:
        company_col_idx = headers.index("Company Name") + 1  # 1-based
    except ValueError:
        company_col_idx = 1  # Fall back to column A

    incoming_name = (row_data.get("Company Name") or "").strip().lower()

    # Look for an existing row matching the company name (skip header row 1)
    target_row_idx: Optional[int] = None
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=company_col_idx).value or ""
        if str(cell_val).strip().lower() == incoming_name:
            target_row_idx = row_idx
            break

    # If no existing row found, append after the last data row
    if target_row_idx is None:
        target_row_idx = ws.max_row + 1

    # Styling helpers
    is_alt = target_row_idx % 2 == 0
    bg = ROW_ALT if is_alt else ROW_NORMAL
    enterprise_type = row_data.get("Enterprise Type", "")
    etype_color = ENTERPRISE_COLORS.get(enterprise_type)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write (or overwrite) every column in the target row
    for col_idx, header in enumerate(headers, 1):
        if not header:
            continue
        value = row_data.get(header, "")
        cell = ws.cell(row=target_row_idx, column=col_idx, value=str(value) if value else "")
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border

        if header == "Enterprise Type" and etype_color:
            cell.fill = PatternFill(fill_type="solid", fgColor=etype_color)
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        elif header == "Pega Usage Confirmed":
            if value == "Yes":
                cell.fill = PatternFill(fill_type="solid", fgColor="D1FAE5")
            elif value == "No":
                cell.fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
            elif value == "Unconfirmed":
                cell.fill = PatternFill(fill_type="solid", fgColor="FEF9C3")
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        else:
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)

    ws.row_dimensions[target_row_idx].height = 30
    wb.save(output_path)
    return output_path
